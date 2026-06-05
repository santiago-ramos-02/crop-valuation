from __future__ import annotations

import argparse
import json
import re
import unicodedata
import warnings
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT_DIR / "20260603Aplicativo.xlsx"
DEFAULT_OUTPUT_DIR = ROOT_DIR / "scripts" / "generated"

BASE_SHEET = "Base_datActividades_priorFIN"
PRICES_SHEET = "Tabla_Costos_Insumos"
AGRONOMIC_SHEET = "Caracteristicas_agronomicas"
YIELD_CURVE_SHEET = "Rendimientos estandar_"
LOOKUPS_SHEET = "Caracteristicas"
LOCATION_SHEET = "departamento"
RATES_SHEET = "Entradas_Tasas"

STAGES_BY_NORMALIZED_NAME = {
    "establecimiento": "establecimiento",
    "improductivo": "improductivo",
    "mantenimiento": "mantenimiento",
    "salvamento": "salvamento",
}


@dataclass
class Catalog:
    departamentos: OrderedDict[str, dict[str, Any]] = field(default_factory=OrderedDict)
    municipios: OrderedDict[str, dict[str, Any]] = field(default_factory=OrderedDict)
    crops: OrderedDict[str, dict[str, Any]] = field(default_factory=OrderedDict)
    varieties: OrderedDict[str, dict[str, Any]] = field(default_factory=OrderedDict)
    profiles: list[dict[str, Any]] = field(default_factory=list)
    yield_curve_points: list[dict[str, Any]] = field(default_factory=list)
    lookup_options: OrderedDict[tuple[str, str], dict[str, Any]] = field(default_factory=OrderedDict)
    cost_template_lines: list[dict[str, Any]] = field(default_factory=list)
    input_price_rows: list[dict[str, Any]] = field(default_factory=list)
    profile_by_crop_variety: dict[tuple[str, str], tuple[str, str]] = field(default_factory=dict)
    profile_by_crop_variety_name: dict[str, tuple[str, str]] = field(default_factory=dict)
    validation: dict[str, Any] = field(
        default_factory=lambda: {
            "missing_cost_template_crop_variety": [],
            "missing_cost_template_stages": [],
            "missing_yield_curve_crop_variety": [],
            "missing_yield_curve_stages": [],
            "missing_price_departments": [],
            "skipped_cost_template_rows": [],
            "skipped_yield_curve_rows": [],
            "skipped_input_price_rows": [],
        }
    )


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def display_text(value: Any) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    if text.isupper():
        return text.title()
    return text


def normalize_text(value: Any) -> str:
    text = clean_text(value)
    if text is None:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.casefold()
    text = re.sub(r"[^\w%]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def slugify(value: Any) -> str:
    normalized = normalize_text(value)
    slug = re.sub(r"[^a-z0-9%]+", "_", normalized)
    slug = slug.strip("_").replace("%", "pct")
    return slug or "sin_valor"


def decimal_or_none(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = clean_text(value)
    if text is None:
        return None
    text = text.replace("$", "").replace(" ", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def jsonable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def raw_excel_row(headers: list[str], values: Iterable[Any]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for idx, value in enumerate(values, start=1):
        header = clean_text(headers[idx - 1]) if idx - 1 < len(headers) else None
        key = header or f"column_{idx}"
        raw[key] = jsonable(value)
    return raw


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(Decimal(str(value)), "f")
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, sort_keys=True)
        return "'" + text.replace("'", "''") + "'::jsonb"
    text = str(value)
    return "'" + text.replace("'", "''") + "'"


def build_insert_sql(
    table: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    conflict_columns: list[str],
    *,
    chunk_size: int = 500,
) -> str:
    if not rows:
        return f"-- No rows generated for public.{table}.\n"

    update_columns = [column for column in columns if column not in conflict_columns]
    conflict_target = ", ".join(conflict_columns)
    assignment_sql = ", ".join(f"{column} = EXCLUDED.{column}" for column in update_columns)
    statements: list[str] = []

    for start in range(0, len(rows), chunk_size):
        chunk = rows[start : start + chunk_size]
        value_lines = []
        for row in chunk:
            values = ", ".join(sql_literal(row.get(column)) for column in columns)
            value_lines.append(f"  ({values})")
        conflict_sql = (
            f"ON CONFLICT ({conflict_target}) DO UPDATE SET {assignment_sql};"
            if assignment_sql
            else f"ON CONFLICT ({conflict_target}) DO NOTHING;"
        )
        statements.append(
            f"INSERT INTO public.{table} ({', '.join(columns)}) VALUES\n"
            + ",\n".join(value_lines)
            + f"\n{conflict_sql}\n"
        )

    return "\n".join(statements)


def write_sql_file(
    output_dir: Path,
    filename: str,
    sections: list[str],
    *,
    source_workbook: Path,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    content = [
        "-- Generated by scripts/import_excel_seed_data.py.",
        f"-- Source workbook: {source_workbook.name}",
        "-- Re-run the importer instead of editing this file by hand.",
        "",
    ]
    content.extend(sections)
    (output_dir / filename).write_text("\n".join(content).rstrip() + "\n", encoding="utf-8")


def add_lookup(
    catalog: Catalog,
    group_key: str,
    value: Any,
    *,
    label: Any | None = None,
    metadata: dict[str, Any] | None = None,
    line_order: int | None = None,
) -> None:
    clean_value = clean_text(value)
    if clean_value is None:
        return
    key = (group_key, normalize_text(clean_value))
    if key in catalog.lookup_options:
        return
    catalog.lookup_options[key] = {
        "group_key": group_key,
        "value": clean_value,
        "label": clean_text(label) or clean_value,
        "metadata": metadata or {},
        "line_order": line_order,
        "active": True,
    }


def merge_lookup_metadata(catalog: Catalog, group_key: str, value: Any, metadata: dict[str, Any]) -> None:
    clean_value = clean_text(value)
    if clean_value is None:
        return
    key = (group_key, normalize_text(clean_value))
    option = catalog.lookup_options.get(key)
    if option is None:
        add_lookup(catalog, group_key, clean_value, metadata=metadata)
        return
    option["metadata"] = {**(option.get("metadata") or {}), **metadata}


def read_locations(workbook: Any, catalog: Catalog) -> None:
    ws = workbook[LOCATION_SHEET]
    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        departamento_name = display_text(row[0] if len(row) > 0 else None)
        municipio_name = display_text(row[1] if len(row) > 1 else None)
        if not departamento_name or not municipio_name:
            continue

        departamento_id = slugify(departamento_name)
        if departamento_id not in catalog.departamentos:
            catalog.departamentos[departamento_id] = {
                "id": departamento_id,
                "name": departamento_name,
                "normalized_name": normalize_text(departamento_name),
                "active": True,
            }

        municipio_slug = slugify(municipio_name)
        municipio_id = f"{departamento_id}_{municipio_slug}"
        if municipio_id in catalog.municipios:
            continue

        catalog.municipios[municipio_id] = {
            "id": municipio_id,
            "departamento_id": departamento_id,
            "name": municipio_name,
            "normalized_name": normalize_text(municipio_name),
            "type": clean_text(row[2] if len(row) > 2 else None),
            "longitude": decimal_or_none(row[3] if len(row) > 3 else None),
            "latitude": decimal_or_none(row[4] if len(row) > 4 else None),
            "active": True,
            "_source_row": source_row,
        }


def register_crop_variety(catalog: Catalog, crop_name: Any, variety_name: Any) -> tuple[str, str] | None:
    clean_crop = display_text(crop_name)
    clean_variety = display_text(variety_name)
    if not clean_crop or not clean_variety:
        return None

    crop_id = slugify(clean_crop)
    variety_id = f"{crop_id}_{slugify(clean_variety)}"

    if crop_id not in catalog.crops:
        catalog.crops[crop_id] = {
            "id": crop_id,
            "name": clean_crop,
            "normalized_name": normalize_text(clean_crop),
            "active": True,
        }

    if variety_id not in catalog.varieties:
        catalog.varieties[variety_id] = {
            "id": variety_id,
            "crop_id": crop_id,
            "name": clean_variety,
            "normalized_name": normalize_text(clean_variety),
            "active": True,
        }

    catalog.profile_by_crop_variety[(normalize_text(clean_crop), normalize_text(clean_variety))] = (
        crop_id,
        variety_id,
    )
    return crop_id, variety_id


def read_agronomic_profiles(workbook: Any, catalog: Catalog) -> None:
    ws = workbook[AGRONOMIC_SHEET]
    headers = [clean_text(cell) or f"column_{idx}" for idx, cell in enumerate(next(ws.iter_rows(max_row=1, values_only=True)), start=1)]

    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and value != "" for value in row):
            continue

        ids = register_crop_variety(catalog, row[0], row[1])
        if ids is None:
            continue
        crop_id, variety_id = ids
        crop_variety_name = clean_text(row[2]) or f"{catalog.crops[crop_id]['name']} {catalog.varieties[variety_id]['name']}"

        catalog.profiles.append(
            {
                "crop_id": crop_id,
                "variety_id": variety_id,
                "crop_variety_name": crop_variety_name,
                "lifecycle_months": decimal_or_none(row[3]),
                "lifecycle_years": decimal_or_none(row[4]),
                "harvest_start_month": decimal_or_none(row[5]),
                "harvest_start_year": decimal_or_none(row[6]),
                "default_row_distance_m": decimal_or_none(row[7]),
                "default_plant_distance_m": decimal_or_none(row[8]),
                "default_density_plants_ha": decimal_or_none(row[9]),
                "harvest_years": decimal_or_none(row[10]),
                "support_years": decimal_or_none(row[11]),
                "source_row": source_row,
                "raw_excel_row": raw_excel_row(headers, row[: len(headers)]),
            }
        )
        catalog.profile_by_crop_variety_name[normalize_text(crop_variety_name)] = (crop_id, variety_id)


def read_lookup_options(workbook: Any, catalog: Catalog) -> None:
    ws = workbook[LOOKUPS_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    # Climate, production system, crop type, stages, planting frames, and fitosanitary condition.
    for idx, row in enumerate(rows[1:9], start=2):
        if normalize_text(row[0] if len(row) > 0 else None) != "columna1":
            add_lookup(
                catalog,
                "climate_type",
                row[0] if len(row) > 0 else None,
                metadata={
                    "altitude_range": clean_text(row[1] if len(row) > 1 else None),
                    "temperature_range": clean_text(row[2] if len(row) > 2 else None),
                    "source_sheet": LOOKUPS_SHEET,
                    "source_row": idx,
                },
                line_order=idx,
            )
        add_lookup(catalog, "production_system", row[4] if len(row) > 4 else None, line_order=idx)
        add_lookup(catalog, "crop_type", row[6] if len(row) > 6 else None, line_order=idx)
        add_lookup(catalog, "production_stage", row[10] if len(row) > 10 else None, line_order=idx)
        add_lookup(catalog, "planting_frame", row[12] if len(row) > 12 else None, line_order=idx)
        add_lookup(catalog, "planting_frame", row[24] if len(row) > 24 else None, line_order=idx)
        fitosanitary_condition = row[26] if len(row) > 26 else None
        normalized_fitosanitary = normalize_text(fitosanitary_condition)
        if normalized_fitosanitary == "bueno":
            fitosanitary_condition = "Buena"
        elif normalized_fitosanitary == "malo":
            fitosanitary_condition = "Mala"

        add_lookup(
            catalog,
            "fitosanitary_condition",
            fitosanitary_condition,
            metadata={
                "severity_range": clean_text(row[27] if len(row) > 27 else None),
                "physiological_impact": clean_text(row[28] if len(row) > 28 else None),
                "valuation_factor": clean_text(row[29] if len(row) > 29 else None),
                "source_sheet": LOOKUPS_SHEET,
                "source_row": idx,
            },
            line_order=idx,
        )

    # Agrologic classes.
    for idx, row in enumerate(rows[1:9], start=2):
        add_lookup(
            catalog,
            "agrologic_class",
            row[15] if len(row) > 15 else None,
            label=row[15] if len(row) > 15 else None,
            metadata={
                "description": clean_text(row[16] if len(row) > 16 else None),
                "slope": clean_text(row[17] if len(row) > 17 else None),
                "slope_description": clean_text(row[18] if len(row) > 18 else None),
                "limitations": clean_text(row[19] if len(row) > 19 else None),
                "mechanization": clean_text(row[20] if len(row) > 20 else None),
                "source_sheet": LOOKUPS_SHEET,
                "source_row": idx,
            },
            line_order=idx,
        )

    # Precipitation and rainfall regime block starts at row 10.
    for idx, row in enumerate(rows[10:16], start=11):
        add_lookup(
            catalog,
            "precipitation_range",
            row[0] if len(row) > 0 else None,
            metadata={
                "technical_classification": clean_text(row[1] if len(row) > 1 else None),
                "agrologic_characterization": clean_text(row[2] if len(row) > 2 else None),
                "source": clean_text(row[3] if len(row) > 3 else None),
                "source_sheet": LOOKUPS_SHEET,
                "source_row": idx,
            },
            line_order=idx,
        )
        add_lookup(
            catalog,
            "rainfall_regime",
            row[5] if len(row) > 5 else None,
            metadata={
                "annual_distribution": clean_text(row[6] if len(row) > 6 else None),
                "suggested_region": clean_text(row[7] if len(row) > 7 else None),
                "source": clean_text(row[8] if len(row) > 8 else None),
                "source_sheet": LOOKUPS_SHEET,
                "source_row": idx,
            },
            line_order=idx,
        )

    for line_order, value in enumerate(
        [
            "Sin disponibilidad de agua",
            "Con disponibilidad de agua estacional",
            "Con disponibilidad de agua permanente",
            "Tiene riego",
        ],
        start=1,
    ):
        add_lookup(catalog, "water_availability", value, line_order=line_order)

    fitosanitary_factors = [
        ("Buena", Decimal("0.95")),
        ("Aceptable", Decimal("0.70")),
        ("Regular", Decimal("0.475")),
        ("Mala", Decimal("0.20")),
    ]
    for line_order, (value, factor) in enumerate(fitosanitary_factors, start=1):
        add_lookup(
            catalog,
            "fitosanitary_condition",
            value,
            metadata={
                "valuation_factor": str(factor),
                "source_sheet": "Formulario",
                "source_cell": "B24",
            },
            line_order=line_order,
        )
        merge_lookup_metadata(
            catalog,
            "fitosanitary_condition",
            value,
            {
                "valuation_factor": str(factor),
                "source_sheet": "Formulario",
                "source_cell": "B24",
            },
        )


def read_discount_rate_options(workbook: Any, catalog: Catalog) -> None:
    ws = workbook[RATES_SHEET]
    for line_order, source_row in enumerate(range(52, 56), start=1):
        method = clean_text(ws.cell(source_row, 2).value)
        rate = decimal_or_none(ws.cell(source_row, 3).value)
        if not method or rate is None:
            continue
        add_lookup(
            catalog,
            "discount_rate_method",
            method,
            metadata={
                "rate_ea": str(rate),
                "interpretation": clean_text(ws.cell(source_row, 4).value),
                "source_sheet": RATES_SHEET,
                "source_row": source_row,
                "default": method == clean_text(ws["C57"].value),
            },
            line_order=line_order,
        )


def find_crop_variety(catalog: Catalog, crop_name: Any, variety_name: Any) -> tuple[str, str] | None:
    return catalog.profile_by_crop_variety.get((normalize_text(crop_name), normalize_text(variety_name)))


def find_crop_variety_by_name(catalog: Catalog, crop_variety_name: Any) -> tuple[str, str] | None:
    return catalog.profile_by_crop_variety_name.get(normalize_text(crop_variety_name))


def line_kind_for(rubro_code: Any, rubro_name: Any, input_name: Any, presentation: Any) -> str:
    if clean_text(rubro_code) == "2." or normalize_text(rubro_name) == "insumos":
        return "input"
    if normalize_text(input_name) == "mano de obra" or normalize_text(presentation) == "jornal":
        return "labor"
    return "other_cost"


def read_cost_template_lines(workbook: Any, catalog: Catalog) -> None:
    ws = workbook[BASE_SHEET]
    rows = ws.iter_rows(values_only=True)
    headers: list[str] = []
    for source_row, row in enumerate(rows, start=1):
        if source_row == 16:
            headers = [clean_text(value) or f"column_{idx}" for idx, value in enumerate(row[:24], start=1)]
            break

    for source_row, row in enumerate(ws.iter_rows(min_row=17, values_only=True), start=17):
        values = list(row[:24])
        if not any(value is not None and value != "" for value in values):
            continue

        crop_name, variety_name, stage_name = values[0], values[1], values[2]
        crop_variety_ids = find_crop_variety(catalog, crop_name, variety_name)
        if crop_variety_ids is None:
            catalog.validation["missing_cost_template_crop_variety"].append(
                {
                    "source_row": source_row,
                    "crop": clean_text(crop_name),
                    "variety": clean_text(variety_name),
                }
            )
            catalog.validation["skipped_cost_template_rows"].append(source_row)
            continue

        stage_id = STAGES_BY_NORMALIZED_NAME.get(normalize_text(stage_name))
        if stage_id is None:
            catalog.validation["missing_cost_template_stages"].append(
                {"source_row": source_row, "stage": clean_text(stage_name)}
            )
            catalog.validation["skipped_cost_template_rows"].append(source_row)
            continue

        kind = line_kind_for(values[3], values[4], values[10], values[11])
        unit_price_mode = {
            "input": "input_price_lookup",
            "labor": "jornal_lookup",
            "other_cost": "fixed",
        }[kind]

        crop_id, variety_id = crop_variety_ids
        input_name = clean_text(values[10])
        catalog.cost_template_lines.append(
            {
                "crop_id": crop_id,
                "variety_id": variety_id,
                "stage_id": stage_id,
                "line_order": source_row - 16,
                "rubro_code": clean_text(values[3]),
                "rubro_name": clean_text(values[4]),
                "subrubro_code": clean_text(values[5]),
                "subrubro_name": clean_text(values[6]),
                "activity_code": clean_text(values[7]),
                "activity_name": clean_text(values[8]),
                "line_kind": kind,
                "input_group_name": clean_text(values[9]),
                "input_name": input_name,
                "normalized_input_name": normalize_text(input_name) if input_name else None,
                "presentation": clean_text(values[11]),
                "quantity": decimal_or_none(values[12]),
                "fixed_unit_price_cop": decimal_or_none(values[13]),
                "unit_price_mode": unit_price_mode,
                "source_sheet": BASE_SHEET,
                "source_row": source_row,
                "raw_excel_row": raw_excel_row(headers, values),
            }
        )


def read_yield_curve_points(workbook: Any, catalog: Catalog) -> None:
    ws = workbook[YIELD_CURVE_SHEET]
    headers = [
        clean_text(value) or f"column_{idx}"
        for idx, value in enumerate(next(ws.iter_rows(max_row=1, values_only=True))[:8], start=1)
    ]

    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[:8])
        age_years = decimal_or_none(values[0])
        crop_variety_name = clean_text(values[1])
        stage_name = clean_text(values[2])
        if age_years is None or not crop_variety_name or not stage_name:
            catalog.validation["skipped_yield_curve_rows"].append(source_row)
            continue

        crop_variety_ids = find_crop_variety_by_name(catalog, crop_variety_name)
        if crop_variety_ids is None:
            catalog.validation["missing_yield_curve_crop_variety"].append(
                {"source_row": source_row, "crop_variety": crop_variety_name}
            )
            catalog.validation["skipped_yield_curve_rows"].append(source_row)
            continue

        stage_id = STAGES_BY_NORMALIZED_NAME.get(normalize_text(stage_name))
        if stage_id is None:
            catalog.validation["missing_yield_curve_stages"].append(
                {"source_row": source_row, "stage": stage_name}
            )
            catalog.validation["skipped_yield_curve_rows"].append(source_row)
            continue

        crop_id, variety_id = crop_variety_ids
        catalog.yield_curve_points.append(
            {
                "crop_id": crop_id,
                "variety_id": variety_id,
                "age_years": age_years,
                "stage_id": stage_id,
                "potential_yield_kg_ha": decimal_or_none(values[3]),
                "default_density_plants_ha": decimal_or_none(values[4]),
                "density_factor": decimal_or_none(values[5]),
                "water_factor": decimal_or_none(values[6]),
                "source_sheet": YIELD_CURVE_SHEET,
                "source_row": source_row,
                "raw_excel_row": raw_excel_row(headers, values),
            }
        )


def read_input_price_rows(workbook: Any, catalog: Catalog) -> None:
    ws = workbook[PRICES_SHEET]
    headers = [
        clean_text(value) or f"column_{idx}"
        for idx, value in enumerate(next(ws.iter_rows(max_row=1, values_only=True))[:10], start=1)
    ]
    departamento_by_normalized = {
        row["normalized_name"]: departamento_id for departamento_id, row in catalog.departamentos.items()
    }

    for source_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[:10])
        departamento_name = clean_text(values[0])
        input_name = clean_text(values[2])
        if not departamento_name or not input_name:
            catalog.validation["skipped_input_price_rows"].append(source_row)
            continue
        if normalize_text(departamento_name) == "total general":
            catalog.validation["skipped_input_price_rows"].append(source_row)
            continue

        departamento_id = departamento_by_normalized.get(normalize_text(departamento_name))
        if departamento_id is None:
            catalog.validation["missing_price_departments"].append(
                {"source_row": source_row, "department": departamento_name}
            )

        catalog.input_price_rows.append(
            {
                "departamento_id": departamento_id,
                "departamento_name_excel": departamento_name,
                "input_group_name": clean_text(values[1]),
                "input_name": input_name,
                "normalized_input_name": normalize_text(input_name),
                "presentation": clean_text(values[3]),
                "average_price_final_cop": decimal_or_none(values[4]),
                "price_source": clean_text(values[5]),
                "expert_price_cop": decimal_or_none(values[6]),
                "region_name": clean_text(values[7]),
                "regional_imputed_price_cop": decimal_or_none(values[8]),
                "calculated_final_price_cop": decimal_or_none(values[9]),
                "source_sheet": PRICES_SHEET,
                "source_row": source_row,
                "raw_excel_row": raw_excel_row(headers, values),
            }
        )


def summarize_naranja_example(catalog: Catalog) -> dict[str, Any]:
    ids = find_crop_variety(catalog, "Naranja", "Común")
    if ids is None:
        return {"found": False}
    crop_id, variety_id = ids
    names = [
        row["input_name"]
        for row in catalog.cost_template_lines
        if row["crop_id"] == crop_id
        and row["variety_id"] == variety_id
        and row["stage_id"] == "mantenimiento"
        and row["line_kind"] == "input"
    ]
    return {"found": True, "input_count": len(names), "input_names": names}


def write_generated_sql(catalog: Catalog, output_dir: Path, workbook_path: Path) -> None:
    write_sql_file(
        output_dir,
        "001_seed_departamentos.sql",
        [
            build_insert_sql(
                "departamentos",
                ["id", "name", "normalized_name", "active"],
                list(catalog.departamentos.values()),
                ["id"],
            )
        ],
        source_workbook=workbook_path,
    )

    municipio_rows = [
        {key: value for key, value in row.items() if not key.startswith("_")}
        for row in catalog.municipios.values()
    ]
    write_sql_file(
        output_dir,
        "002_seed_municipios.sql",
        [
            build_insert_sql(
                "municipios",
                [
                    "id",
                    "departamento_id",
                    "name",
                    "normalized_name",
                    "type",
                    "longitude",
                    "latitude",
                    "active",
                ],
                municipio_rows,
                ["id"],
            )
        ],
        source_workbook=workbook_path,
    )

    write_sql_file(
        output_dir,
        "003_seed_lookup_options.sql",
        [
            build_insert_sql(
                "lookup_options",
                ["group_key", "value", "label", "metadata", "line_order", "active"],
                list(catalog.lookup_options.values()),
                ["group_key", "value"],
            )
        ],
        source_workbook=workbook_path,
    )

    write_sql_file(
        output_dir,
        "004_seed_crops_varieties_profiles.sql",
        [
            build_insert_sql(
                "crops",
                ["id", "name", "normalized_name", "active"],
                list(catalog.crops.values()),
                ["id"],
            ),
            build_insert_sql(
                "varieties",
                ["id", "crop_id", "name", "normalized_name", "active"],
                list(catalog.varieties.values()),
                ["id"],
            ),
            build_insert_sql(
                "crop_variety_agronomic_profiles",
                [
                    "crop_id",
                    "variety_id",
                    "crop_variety_name",
                    "lifecycle_months",
                    "lifecycle_years",
                    "harvest_start_month",
                    "harvest_start_year",
                    "default_row_distance_m",
                    "default_plant_distance_m",
                    "default_density_plants_ha",
                    "harvest_years",
                    "support_years",
                    "source_row",
                    "raw_excel_row",
                ],
                catalog.profiles,
                ["crop_id", "variety_id"],
            ),
            build_insert_sql(
                "yield_curve_points",
                [
                    "crop_id",
                    "variety_id",
                    "age_years",
                    "stage_id",
                    "potential_yield_kg_ha",
                    "default_density_plants_ha",
                    "density_factor",
                    "water_factor",
                    "source_sheet",
                    "source_row",
                    "raw_excel_row",
                ],
                catalog.yield_curve_points,
                ["source_sheet", "source_row"],
            ),
        ],
        source_workbook=workbook_path,
    )

    write_sql_file(
        output_dir,
        "005_seed_cost_template_lines.sql",
        [
            build_insert_sql(
                "cost_template_lines",
                [
                    "crop_id",
                    "variety_id",
                    "stage_id",
                    "line_order",
                    "rubro_code",
                    "rubro_name",
                    "subrubro_code",
                    "subrubro_name",
                    "activity_code",
                    "activity_name",
                    "line_kind",
                    "input_group_name",
                    "input_name",
                    "normalized_input_name",
                    "presentation",
                    "quantity",
                    "fixed_unit_price_cop",
                    "unit_price_mode",
                    "source_sheet",
                    "source_row",
                    "raw_excel_row",
                ],
                catalog.cost_template_lines,
                ["source_sheet", "source_row"],
            )
        ],
        source_workbook=workbook_path,
    )

    write_sql_file(
        output_dir,
        "006_seed_input_price_rows.sql",
        [
            build_insert_sql(
                "input_price_rows",
                [
                    "departamento_id",
                    "departamento_name_excel",
                    "input_group_name",
                    "input_name",
                    "normalized_input_name",
                    "presentation",
                    "average_price_final_cop",
                    "price_source",
                    "expert_price_cop",
                    "region_name",
                    "regional_imputed_price_cop",
                    "calculated_final_price_cop",
                    "source_sheet",
                    "source_row",
                    "raw_excel_row",
                ],
                catalog.input_price_rows,
                ["source_sheet", "source_row"],
            )
        ],
        source_workbook=workbook_path,
    )


def write_validation_report(catalog: Catalog, output_dir: Path) -> dict[str, Any]:
    missing_price_departments = sorted(
        {entry["department"] for entry in catalog.validation["missing_price_departments"]}
    )
    missing_template_refs = [
        dict(item)
        for item in {
            (entry["crop"], entry["variety"], entry["source_row"]): entry
            for entry in catalog.validation["missing_cost_template_crop_variety"]
        }.values()
    ]

    report = {
        "counts": {
            "departamentos": len(catalog.departamentos),
            "municipios": len(catalog.municipios),
            "lookup_options": len(catalog.lookup_options),
            "crops": len(catalog.crops),
            "varieties": len(catalog.varieties),
            "agronomic_profiles": len(catalog.profiles),
            "yield_curve_points": len(catalog.yield_curve_points),
            "cost_template_lines": len(catalog.cost_template_lines),
            "input_price_rows": len(catalog.input_price_rows),
            "input_cost_template_lines": sum(
                1 for row in catalog.cost_template_lines if row["line_kind"] == "input"
            ),
        },
        "missing_references": {
            "cost_template_crop_variety": missing_template_refs,
            "cost_template_stages": catalog.validation["missing_cost_template_stages"],
            "yield_curve_crop_variety": catalog.validation["missing_yield_curve_crop_variety"],
            "yield_curve_stages": catalog.validation["missing_yield_curve_stages"],
            "price_departments": missing_price_departments,
        },
        "skipped_rows": {
            "cost_template_rows": catalog.validation["skipped_cost_template_rows"],
            "yield_curve_rows": catalog.validation["skipped_yield_curve_rows"],
            "input_price_rows": catalog.validation["skipped_input_price_rows"],
        },
        "known_examples": {
            "naranja_comun_mantenimiento": summarize_naranja_example(catalog),
        },
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "seed_validation_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return report


def build_catalog(workbook_path: Path) -> Catalog:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    required_sheets = {
        LOCATION_SHEET,
        LOOKUPS_SHEET,
        AGRONOMIC_SHEET,
        YIELD_CURVE_SHEET,
        BASE_SHEET,
        PRICES_SHEET,
        RATES_SHEET,
    }
    missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
    if missing_sheets:
        raise RuntimeError(f"Workbook is missing required sheets: {', '.join(missing_sheets)}")

    catalog = Catalog()
    read_locations(workbook, catalog)
    read_agronomic_profiles(workbook, catalog)
    read_lookup_options(workbook, catalog)
    read_discount_rate_options(workbook, catalog)
    read_yield_curve_points(workbook, catalog)
    read_cost_template_lines(workbook, catalog)
    read_input_price_rows(workbook, catalog)
    return catalog


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate seed SQL from 20260603Aplicativo.xlsx.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    output_dir = args.output_dir.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    catalog = build_catalog(workbook_path)
    write_generated_sql(catalog, output_dir, workbook_path)
    report = write_validation_report(catalog, output_dir)

    print(f"Generated seed SQL in {output_dir}")
    for key, value in report["counts"].items():
        print(f"{key}: {value}")
    if report["missing_references"]["price_departments"]:
        print("Missing price department mappings:")
        for department in report["missing_references"]["price_departments"]:
            print(f"  - {department}")
    if report["missing_references"]["cost_template_crop_variety"]:
        print("Missing cost template crop/variety mappings:")
        for entry in report["missing_references"]["cost_template_crop_variety"]:
            print(f"  - row {entry['source_row']}: {entry['crop']} / {entry['variety']}")


if __name__ == "__main__":
    main()
